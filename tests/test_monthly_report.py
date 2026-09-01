import importlib.util
import io
import json
import os
import sys
import unittest
import zipfile
from xml.etree import ElementTree as ET
from datetime import datetime, timezone
from decimal import Decimal
from unittest.mock import Mock, patch

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(__file__))
SPEC = importlib.util.spec_from_file_location("monthly_report", os.path.join(ROOT, "lambda", "monthly_report", "index.py"))
monthly = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = monthly
SPEC.loader.exec_module(monthly)
BACKFILL_SPEC = importlib.util.spec_from_file_location("backfill_monthly_reports", os.path.join(ROOT, "scripts", "backfill_monthly_reports.py"))
backfill = importlib.util.module_from_spec(BACKFILL_SPEC)
sys.modules[BACKFILL_SPEC.name] = backfill
BACKFILL_SPEC.loader.exec_module(backfill)


class HttpResponse:
    def __init__(self, payload, status=200):
        self.payload = payload
        self.status = status
    def __enter__(self): return self
    def __exit__(self, *args): return False
    def read(self): return json.dumps(self.payload).encode("utf-8")


class Body:
    def __init__(self, value): self.value = value
    def read(self): return self.value


class MissingKey(Exception):
    def __init__(self): self.response = {"Error": {"Code": "NoSuchKey"}}


class MonthlyReportTests(unittest.TestCase):
    def test_month_selection_and_notification_switches(self):
        selected = monthly.select_report_month({"time": "2026-01-02T06:00:00Z", "report_type": "final"})
        self.assertEqual("2025-12", selected["month"])
        self.assertFalse(selected["notify"])
        enabled = monthly.select_report_month({
            "time": "2026-01-02T06:00:00Z", "report_type": "final", "notify": True,
        })
        self.assertTrue(enabled["notify"])
        disabled = monthly.select_report_month({
            "time": "2026-01-02T06:00:00Z", "report_type": "final", "notify": False,
        })
        self.assertFalse(disabled["notify"])
        provisional = monthly.select_report_month({"time": "2026-03-01T06:00:00Z", "report_type": "provisional"})
        self.assertEqual(("2026-02", False), (provisional["month"], provisional["notify"]))
        manual = monthly.select_report_month({"month": "2026-04", "notify": "false"})
        self.assertFalse(manual["scheduled"])
        self.assertFalse(manual["notify"])
        self.assertEqual(["dev"], monthly.select_notification_channels({}))
        self.assertEqual(["prod"], monthly.select_notification_channels({"notification_channel": "prod"}))
        self.assertEqual(["dev", "prod"], monthly.select_notification_channels({"notification_channel": "both"}))
        with self.assertRaises(ValueError):
            monthly.select_notification_channels({"notification_channel": "production"})
        with self.assertRaises(ValueError):
            monthly.select_report_month({"month": "2026-04", "notify": "sometimes"})

    def test_february_partial(self):
        self.assertEqual("PARTIAL", monthly.report_status("2026-02")[0])
        self.assertIn("2026-02-10", monthly.report_status("2026-02")[1])
        self.assertEqual("COMPLETE", monthly.report_status("2026-03")[0])

    def test_tier_normalization_and_precedence(self):
        cases = {"PRO_PLUS": "Pro+", "Kiro-Pro+-users": "Pro+", "pro max": "Pro Max", "PRO_MAX": "Pro Max", "Power users": "Power", "Kiro-Pro-users": "Pro"}
        for raw, expected in cases.items():
            self.assertEqual(expected, monthly.normalize_tier(raw), raw)
        self.assertEqual("Power", monthly.highest_tier(["Pro", "Pro+", "Pro Max", "Power"]))

    def test_formula_injection_and_illegal_chars(self):
        self.assertEqual("'=2+2 bad", monthly.sanitize("=2+2\x00\nbad"))
        self.assertEqual("'+SUM(A1)", monthly.sanitize("+SUM(A1)"))

    def test_deterministic_keys(self):
        expected = "dashboard-reports/public/kiro-monthly/2026/03/kiro-credits-2026-03-final.xlsx"
        self.assertEqual(expected, monthly.monthly_keys("dashboard-reports/public/kiro-monthly", "2026-03", "final")["workbook"])
        self.assertEqual(monthly.monthly_keys("p", "2026-03", "final"), monthly.monthly_keys("p", "2026-03", "final"))

    def test_csv_valid_empty_is_authoritative(self):
        s3 = Mock()
        s3.get_object.return_value = {"Body": Body(b"userid,user_name,subscription_tier\n")}
        roster, authoritative = monthly.load_csv_roster(s3, "bucket", "exact.csv")
        self.assertEqual({}, roster)
        self.assertTrue(authoritative)
        s3.get_object.side_effect = MissingKey()
        roster, authoritative = monthly.load_csv_roster(s3, "bucket", "missing.csv")
        self.assertIsNone(roster)
        self.assertFalse(authoritative)

    def test_csv_alias_duplicate_id_highest_tier(self):
        data = (
            "user_id,name,email,plan\n"
            "d-store.u1,Alice,alice@example.com,Pro\n"
            "u1,u1,,Power\n"
        ).encode("utf-8-sig")
        roster = monthly.parse_subscription_csv(data)
        self.assertEqual("Power", roster["u1"]["subscription_tier"])
        self.assertEqual("Alice", roster["u1"]["user_name"])
        self.assertEqual("alice@example.com", roster["u1"]["email"])

    def test_identity_mapping_is_identity_only(self):
        data = (
            "userid,username,email,subscription_status,subscription_tier\n"
            "d-store.u1,Alice,a@example.com,ACTIVE,Power\n"
            "u1,u1,,ACTIVE,Pro\n"
        ).encode()
        identities = monthly.parse_identity_mapping_csv(data)
        self.assertEqual({"user_name": "Alice", "email": "a@example.com"}, identities["u1"])
        self.assertNotIn("subscription_status", identities["u1"])
        self.assertNotIn("subscription_tier", identities["u1"])

    def test_duplicate_names_are_not_merged(self):
        usage = monthly.usage_rows_by_id([
            {"user_id": "u1", "credits": "10", "latest_tier": "PRO"},
            {"user_id": "u2", "credits": "20", "latest_tier": "PRO"},
        ])
        roster = {
            "u1": {"user_name": "Same", "subscription_tier": "Pro"},
            "u2": {"user_name": "Same", "subscription_tier": "Pro"},
        }
        rows = monthly.merge_report_rows("2026-03", usage, roster, True)
        self.assertEqual(2, len(rows))
        self.assertEqual({"u1", "u2"}, {r["user_id"] for r in rows})

    def test_threshold_rules(self):
        self.assertEqual(("red", ""), monthly.usage_band(0, 1000))
        self.assertEqual("red", monthly.usage_band(99, 1000)[0])
        self.assertEqual("yellow", monthly.usage_band(100, 1000)[0])
        self.assertEqual("green", monthly.usage_band(500, 1000)[0])
        self.assertEqual("capacity_pressure", monthly.usage_band(900, 1000)[1])
        self.assertEqual("exceeded", monthly.usage_band(1000, 1000)[1])

    def test_low_usage_is_ranked_first(self):
        usage = {
            "high": {"credits": "600", "latest_tier": "Pro", "tier_history": ["Pro"]},
            "low_large_plan": {"credits": "200", "latest_tier": "Power", "tier_history": ["Power"]},
            "low_small_plan": {"credits": "20", "latest_tier": "Pro", "tier_history": ["Pro"]},
            "zero": {"credits": "0", "latest_tier": "Pro", "tier_history": ["Pro"]},
        }
        roster = {uid: {"user_name": uid, "subscription_tier": item["latest_tier"]} for uid, item in usage.items()}
        rows = monthly.merge_report_rows("2026-03", usage, roster, True)
        self.assertEqual(["zero", "low_small_plan", "low_large_plan", "high"], [row["user_id"] for row in rows])
        self.assertEqual(sorted(row["usage_rate"] for row in rows), [row["usage_rate"] for row in rows])

    def sample_rows(self):
        return monthly.merge_report_rows("2026-03", {"u1": {"credits": Decimal("120"), "overage": 0, "latest_tier": "Pro", "tier_history": ["Pro"], "active_days": 2}}, {"u1": {"user_name": "Alice", "email": "alice@example.com", "subscription_tier": "Pro", "subscription_status": "ACTIVE", "plan_source": "test"}}, True)

    def test_workbook_sheets_formulas_and_reopen(self):
        data = monthly.build_monthly_workbook(self.sample_rows(), "2026-03", "COMPLETE")
        wb = load_workbook(io.BytesIO(data), data_only=False)
        self.assertEqual(monthly.SHEETS, wb.sheetnames)
        detail = wb["月度用户明细"]
        self.assertEqual("=IFERROR(IF(J2>0,I2/J2,0),0)", detail["K2"].value)
        self.assertEqual("=IFERROR(IF(J2>0,L2*MAX(0,1-K2),0),0)", detail["M2"].value)
        self.assertTrue(wb.calculation.fullCalcOnLoad)
        self.assert_no_duplicate_filter_owners(data)

    def assert_no_duplicate_filter_owners(self, data):
        namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            names = workbook.find("x:definedNames", namespace)
            self.assertTrue(names is None or not any("_FilterDatabase" in item.attrib.get("name", "") for item in names))
            worksheets = [name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]
            tables = [name for name in archive.namelist() if name.startswith("xl/tables/table") and name.endswith(".xml")]
            self.assertLessEqual(len(tables), len(worksheets))
            for name in worksheets:
                root = ET.fromstring(archive.read(name))
                self.assertIsNone(root.find("x:autoFilter", namespace), name)
                table_parts = root.find("x:tableParts", namespace)
                if table_parts is not None:
                    self.assertGreater(int(table_parts.attrib.get("count", "0")), 0, name)
            for name in tables:
                root = ET.fromstring(archive.read(name))
                auto_filter = root.find("x:autoFilter", namespace)
                self.assertIsNotNone(auto_filter, name)
                self.assertEqual(root.attrib["ref"], auto_filter.attrib["ref"], name)

    def test_annual_rebuild_helper_is_idempotent_by_content(self):
        rows = self.sample_rows()
        csv_rows = monthly.csv_to_rows(monthly.rows_to_csv(rows))
        one = monthly.build_annual_workbook({"2026-03": csv_rows}, "2026")
        two = monthly.build_annual_workbook({"2026-03": csv_rows}, "2026")
        wb1 = load_workbook(io.BytesIO(one), data_only=False)
        wb2 = load_workbook(io.BytesIO(two), data_only=False)
        self.assertEqual(["年度汇总", "2026-03"], wb1.sheetnames)
        self.assertEqual(list(wb1["年度汇总"].values), list(wb2["年度汇总"].values))
        self.assert_no_duplicate_filter_owners(one)

    def test_annual_rebuild_preserves_negative_mom_change(self):
        rows = self.sample_rows()
        rows[0]["previous_month_credits"] = Decimal("240")
        rows[0]["mom_change"] = Decimal("-0.5")
        rows[0]["user_name"] = "=unsafe"
        csv_rows = monthly.csv_to_rows(monthly.rows_to_csv(rows))
        self.assertEqual("-0.5", csv_rows[0]["mom_change"])
        self.assertEqual("'=unsafe", csv_rows[0]["user_name"])
        annual = monthly.build_annual_workbook({"2026-03": csv_rows}, "2026")
        workbook = load_workbook(io.BytesIO(annual), data_only=False)
        self.assertEqual(-0.5, workbook["2026-03"]["W2"].value)

    def test_historical_roster_guard(self):
        self.assertTrue(monthly.should_use_current_roster("2026-07", "2026-08"))
        self.assertFalse(monthly.should_use_current_roster("2026-06", "2026-08"))
        historical = monthly.merge_report_rows(
            "2026-02",
            {"u1": {"credits": Decimal("1"), "latest_tier": "Pro", "tier_history": ["Pro"]}},
            {
                "u1": {"user_name": "Current user", "email": "current@example.com", "subscription_tier": "Power", "subscription_status": "ACTIVE", "plan_source": "current", "activation_date": "2026-02-15"},
                "u2": {"user_name": "Identity only", "subscription_tier": "Power", "subscription_status": "ACTIVE"},
            },
            False,
            identity_map={
                "u1": {"user_name": "Mapped user", "email": "mapped@example.com"},
                "u2": {"user_name": "Identity only", "email": "u2@example.com"},
            },
        )
        self.assertEqual(["u1"], [row["user_id"] for row in historical])
        self.assertEqual("", historical[0]["new_subscription"])
        self.assertEqual("", historical[0]["activation_date"])
        self.assertEqual("Mapped user", historical[0]["user_name"])
        self.assertEqual("mapped@example.com", historical[0]["email"])
        self.assertEqual("UNKNOWN", historical[0]["subscription_status"])
        self.assertEqual("usage report", historical[0]["plan_source"])
        self.assertEqual("Pro", historical[0]["month_end_tier"])
        self.assertFalse(monthly.should_write_subscription_snapshot(False))
        self.assertTrue(monthly.should_write_subscription_snapshot(True))

    def test_same_day_tier_selection_is_deterministic_in_sql(self):
        sql = monthly.build_usage_sql("kiro_analytics", "2026-08")
        latest = monthly.build_latest_tiers_sql("kiro_analytics")
        self.assertIn("CASE MAX(tier_rank)", sql)
        self.assertIn("array_agg(tier_label)", sql)
        self.assertNotIn("max_by(subscription_tier, usage_date)", sql)
        self.assertIn("CASE MAX(tier_rank)", latest)

    def card_rows(self):
        rows = []
        low_rates = [Decimal("0.08"), Decimal("0.01"), Decimal("0.09"), Decimal("0.03"), Decimal("0.05")]
        for index in range(32):
            if index < 15:
                color, rate, credits = "red", Decimal("0"), Decimal("0")
            elif index < 20:
                color, rate = "red", low_rates[index - 15]
                credits = rate * Decimal("1000")
            elif index < 22:
                color, rate, credits = "yellow", Decimal("0.3"), Decimal("300")
            else:
                color, rate, credits = "green", Decimal("0.6"), Decimal("600")
            uid = f"user-{index:08d}"
            name = "重名用户" if index in {0, 1} else uid if index == 2 else f"用户{index:02d}"
            cost = Decimal("120") if index == 31 else Decimal("20")
            rows.append({
                "user_id": uid, "user_name": name, "email": f"user{index}@example.com",
                "month_end_tier": "Pro", "credits": credits, "capacity": 1000,
                "usage_rate": rate, "color": color, "estimated_plan_cost": cost,
                "unused_capacity_value": cost * (Decimal("1") - rate),
                "consecutive_zero_months": 3 if index == 0 else 1 if index < 15 else 0,
                "consecutive_low_months": 3 if index == 16 else 2 if 15 < index < 20 else 1,
            })
        return rows

    def card_contents(self, value):
        result = []
        if isinstance(value, dict):
            if isinstance(value.get("content"), str):
                result.append(value["content"])
            for item in value.values():
                result.extend(self.card_contents(item))
        elif isinstance(value, list):
            for item in value:
                result.extend(self.card_contents(item))
        return result

    def card_component_count(self, value):
        if isinstance(value, dict):
            return (1 if "tag" in value else 0) + sum(self.card_component_count(item) for item in value.values())
        if isinstance(value, list):
            return sum(self.card_component_count(item) for item in value)
        return 0

    def find_card_element(self, card, element_id):
        return next(element for element in card["body"]["elements"] if element.get("element_id") == element_id)

    def test_feishu_card_v2_content_layout_and_sorting(self):
        url = "http://example.invalid/report.xlsx"
        payload = monthly.build_feishu_card("2026-07", self.card_rows(), url)
        self.assertEqual("interactive", payload["msg_type"])
        card = payload["card"]
        self.assertEqual("2.0", card["schema"])
        self.assertTrue(card["config"]["update_multi"])
        self.assertEqual("fill", card["config"]["width_mode"])
        self.assertEqual(url, card["card_link"]["url"])
        self.assertEqual("blue", card["header"]["template"])
        self.assertEqual("【2026-07】Kiro订阅用户月度用量分析月报", card["header"]["title"]["content"])
        self.assertEqual("自然月 · Final · 数据完整", card["header"]["subtitle"]["content"])
        self.assertEqual(
            [("red", "零使用 15"), ("orange", "低用 5")],
            [(tag["color"], tag["text"]["content"]) for tag in card["header"]["text_tag_list"]],
        )

        colors = card["config"]["style"]["color"]
        self.assertEqual({"neutral_bg", "panel_bg", "border_soft", "advice_bg"}, set(colors))
        for token in colors.values():
            self.assertEqual({"light_mode", "dark_mode"}, set(token))
            self.assertTrue(token["light_mode"].startswith("rgba("))
            self.assertTrue(token["dark_mode"].startswith("rgba("))
            self.assertNotEqual(token["light_mode"], token["dark_mode"])

        elements = card["body"]["elements"]
        self.assertEqual(["column_set", "column_set", "markdown", "collapsible_panel", "collapsible_panel", "column_set", "button"], [e["tag"] for e in elements])
        for kpi_row in elements[:2]:
            self.assertEqual("bisect", kpi_row["flex_mode"])
            self.assertEqual(2, len(kpi_row["columns"]))
            self.assertTrue(all(column["width"] == "weighted" for column in kpi_row["columns"]))
            self.assertTrue(all(column["background_style"] == "neutral_bg" for column in kpi_row["columns"]))
        kpis = "\n".join(self.card_contents(elements[:3]))
        self.assertIn("👥 订阅用户\n**32 人**", kpis)
        self.assertIn("💰 月度总成本\n**$740.00**", kpis)
        self.assertIn("⛔ 零使用\n**15 人 · 47%**\n订阅成本 $300.00", kpis)
        self.assertIn("⚠️ 非零低用量\n**5 人 · 16%**\n低效容量价值 $94.80", kpis)
        self.assertIn("中用量 **2人（6%）**", kpis)
        self.assertIn("高用量 **10人（31%）**", kpis)

        zero = self.find_card_element(card, "zero_users")
        self.assertTrue(zero["expanded"])
        self.assertEqual("panel_bg", zero["background_color"])
        self.assertEqual("neutral_bg", zero["header"]["background_color"])
        self.assertEqual("border_soft", zero["border"]["color"])
        self.assertEqual("grey", zero["header"]["icon"]["color"])
        zero_rows = [item for item in zero["elements"] if item["tag"] == "column_set"]
        self.assertEqual(15, len(zero_rows))
        self.assertEqual("none", zero_rows[0]["flex_mode"])
        self.assertEqual([3, 2], [column["weight"] for column in zero_rows[0]["columns"]])
        zero_text = "\n".join(self.card_contents(zero))
        self.assertIn("⛔ **零使用用户 · 15人**", zero_text)
        self.assertIn("重名用户（user0@example.com）", zero_text)
        self.assertIn("重名用户（user1@example.com）", zero_text)
        self.assertIn("未识别用户（00000002）", zero_text)
        self.assertIn("连续零使用 **3个月**", self.card_contents(zero_rows[0])[0])
        self.assertNotIn("0 / 1,000", zero_text)

        low = self.find_card_element(card, "low_users")
        self.assertTrue(low["expanded"])
        self.assertEqual("panel_bg", low["background_color"])
        self.assertEqual("neutral_bg", low["header"]["background_color"])
        self.assertEqual("border_soft", low["border"]["color"])
        self.assertEqual("grey", low["header"]["icon"]["color"])
        low_rows = [item for item in low["elements"] if item["tag"] == "column_set"]
        self.assertEqual(5, len(low_rows))
        first_low = "\n".join(self.card_contents(low_rows[0]))
        self.assertIn("⚠️ **非零低用量用户 · 5人**", "\n".join(self.card_contents(low)))
        self.assertIn("用户16", first_low)
        self.assertIn("10 / 1,000 Credits · 1.0%", first_low)
        self.assertIn("连续低用 **3个月**", first_low)
        self.assertNotIn("用户22", zero_text + "\n".join(self.card_contents(low)))

        self.assertEqual("advice_bg", elements[5]["background_style"])
        advice = "\n".join(self.card_contents(elements[5]))
        self.assertIn("💡 **管理建议**", advice)
        self.assertIn("连续零使用 ≥2个月：**1人**", advice)
        self.assertIn("首次/单月零使用：**14人**", advice)
        self.assertIn("连续低用 ≥2个月：**4人**", advice)
        self.assertIn("首次/单月低用：**1人**", advice)
        button = elements[-1]
        self.assertEqual("primary_filled", button["type"])
        self.assertEqual("fill", button["width"])
        self.assertEqual("📊 查看完整 Excel 报告", button["text"]["content"])
        self.assertEqual({"type": "open_url", "default_url": url, "pc_url": url, "ios_url": url, "android_url": url}, button["behaviors"][0])

        all_text = "\n".join(self.card_contents(card))
        for icon in ("👥", "💰", "⛔", "⚠️", "💡", "📊"):
            self.assertIn(icon, all_text)
        for decorative_icon in ("🚨", "🔥", "🎉", "🔴", "🟠"):
            self.assertNotIn(decorative_icon, all_text)
        rendered = json.dumps(card, ensure_ascii=False)
        self.assertNotIn('"background_style": "red"', rendered)
        self.assertNotIn('"background_style": "orange"', rendered)
        self.assertNotIn('"background_color": "red"', rendered)
        self.assertNotIn('"background_color": "orange"', rendered)
        self.assertNotIn("action", [item.get("tag") for item in elements])
        self.assertLessEqual(self.card_component_count(card), 200)
        self.assertLess(len(json.dumps(payload, ensure_ascii=False).encode("utf-8")), 20000)

    def test_feishu_card_v2_dynamic_fold_at_twenty_one_risks(self):
        rows = self.card_rows()
        rows.append({
            "user_id": "extra-low", "user_name": "额外低用用户", "email": "extra@example.com",
            "month_end_tier": "Pro", "credits": 20, "capacity": 1000,
            "usage_rate": Decimal("0.02"), "color": "red", "estimated_plan_cost": 20,
            "unused_capacity_value": Decimal("19.6"), "consecutive_low_months": 1,
            "consecutive_zero_months": 0,
        })
        card = monthly.build_feishu_card("2026-07", rows, "http://example.invalid/report.xlsx")["card"]
        self.assertTrue(self.find_card_element(card, "zero_users")["expanded"])
        low = self.find_card_element(card, "low_users")
        self.assertFalse(low["expanded"])
        self.assertEqual(6, len([item for item in low["elements"] if item["tag"] == "column_set"]))
        self.assertLessEqual(self.card_component_count(card), 200)

    def test_feishu_card_v2_dense_layout_preserves_large_risk_list(self):
        rows = self.card_rows()
        for index, row in enumerate(rows[20:], start=20):
            row.update({
                "credits": 0, "usage_rate": Decimal(0), "color": "red",
                "unused_capacity_value": row["estimated_plan_cost"],
                "consecutive_zero_months": 1 + index % 3,
                "consecutive_low_months": 0,
            })
        payload = monthly.build_feishu_card("2026-07", rows, "http://example.invalid/report.xlsx")
        card = payload["card"]
        zero = self.find_card_element(card, "zero_users")
        low = self.find_card_element(card, "low_users")
        self.assertEqual(["markdown"], [item["tag"] for item in zero["elements"]])
        self.assertEqual(["markdown"], [item["tag"] for item in low["elements"]])
        risk_text = "\n".join(self.card_contents([zero, low]))
        self.assertIn("用户31", risk_text)
        self.assertIn("连续零使用", risk_text)
        self.assertIn("10 / 1,000 Credits · 1.0% · 连续低用", risk_text)
        self.assertLessEqual(self.card_component_count(card), 200)
        self.assertLess(len(json.dumps(payload, ensure_ascii=False).encode("utf-8")), 20000)

    def test_feishu_card_v2_empty_and_partial_branches(self):
        green = [{
            "user_id": "u1", "user_name": "正常用户", "month_end_tier": "Pro",
            "credits": 600, "capacity": 1000, "usage_rate": Decimal("0.6"),
            "color": "green", "estimated_plan_cost": 20, "unused_capacity_value": 8,
        }]
        complete = monthly.build_feishu_card("2026-07", green, "http://example.invalid/report.xlsx")
        self.assertEqual("blue", complete["card"]["header"]["template"])
        self.assertEqual([], complete["card"]["header"]["text_tag_list"])
        text = "\n".join(self.card_contents(complete["card"]["body"]["elements"]))
        self.assertIn("本月无零使用用户", text)
        self.assertIn("本月无非零低用量用户", text)
        self.assertIn("没有需要回收或调整的低效订阅", text)
        partial = monthly.build_feishu_card("2026-08", green, "http://example.invalid/report.xlsx", "PARTIAL", "当前月未结束")
        self.assertEqual("blue", partial["card"]["header"]["template"])
        self.assertTrue(partial["card"]["header"]["title"]["content"].endswith("（部分数据）"))
        self.assertEqual("自然月 · Final · 部分数据", partial["card"]["header"]["subtitle"]["content"])
        partial_text = "\n".join(self.card_contents(partial["card"]["body"]["elements"]))
        self.assertIn("数据说明：", partial_text)
        self.assertIn("当前月未结束", partial_text)

    def test_feishu_card_v2_cross_group_duplicate_and_single_low_advice(self):
        rows = [
            {
                "user_id": "risk-user", "user_name": "同名用户", "email": "risk@example.com",
                "month_end_tier": "Pro", "credits": 50, "capacity": 1000,
                "usage_rate": Decimal("0.05"), "color": "red", "estimated_plan_cost": 20,
                "unused_capacity_value": 19, "consecutive_low_months": 1,
            },
            {
                "user_id": "green-user", "user_name": "同名用户", "email": "green@example.com",
                "month_end_tier": "Pro", "credits": 600, "capacity": 1000,
                "usage_rate": Decimal("0.6"), "color": "green", "estimated_plan_cost": 20,
                "unused_capacity_value": 8, "consecutive_low_months": 0,
            },
        ]
        card = monthly.build_feishu_card("2026-07", rows, "http://example.invalid/report.xlsx")["card"]
        self.assertEqual([("orange", "低用 1")], [(tag["color"], tag["text"]["content"]) for tag in card["header"]["text_tag_list"]])
        low = "\n".join(self.card_contents(self.find_card_element(card, "low_users")))
        advice = "\n".join(self.card_contents(card["body"]["elements"][-2]))
        self.assertIn("同名用户（risk@example.com）", low)
        self.assertIn("首次/单月低用：**1人**", advice)
        self.assertNotIn("没有需要回收或调整的低效订阅", advice)

    def test_feishu_interactive_send_and_channel_routing(self):
        payload = monthly.build_feishu_card("2026-07", self.card_rows(), "http://example.invalid/report.xlsx")
        secrets = Mock()
        secrets.get_secret_value.return_value = {"SecretString": json.dumps({"webhook": "https://example.invalid/hook"})}
        with patch.object(monthly.urllib.request, "urlopen", return_value=HttpResponse({"code": 0})) as request:
            self.assertIsNone(monthly.send_feishu(secrets, "arn:secret", payload))
            sent = json.loads(request.call_args.args[0].data.decode("utf-8"))
            self.assertEqual(payload, sent)
        with patch.object(monthly.urllib.request, "urlopen", return_value=HttpResponse({"code": 19024, "msg": "rejected"})):
            error = monthly.send_feishu(secrets, "arn:secret", payload)
        self.assertEqual("Feishu notification failed: RuntimeError", error)
        self.assertNotIn("example.invalid", error)

        secrets.reset_mock()
        with patch.object(monthly.urllib.request, "urlopen") as request:
            results = monthly.send_feishu_channels(
                secrets, {"dev": "arn:dev", "prod": ""}, ["prod"], payload,
            )
        self.assertEqual({"prod": "Feishu notification channel is not configured"}, results)
        secrets.get_secret_value.assert_not_called()
        request.assert_not_called()

        secrets.reset_mock()
        secrets.get_secret_value.side_effect = lambda SecretId: {
            "SecretString": json.dumps({"webhook": f"https://example.invalid/{SecretId}"}),
        }
        with patch.object(monthly.urllib.request, "urlopen", return_value=HttpResponse({"code": 0})) as request:
            results = monthly.send_feishu_channels(
                secrets, {"dev": "arn:dev", "prod": "arn:prod"}, ["dev", "prod"], payload,
            )
        self.assertEqual({"dev": None, "prod": None}, results)
        self.assertEqual(["arn:dev", "arn:prod"], [call.kwargs["SecretId"] for call in secrets.get_secret_value.call_args_list])
        self.assertEqual(2, request.call_count)

        secrets.reset_mock()
        with patch.object(monthly.urllib.request, "urlopen") as request:
            results = monthly.send_feishu_channels(
                secrets, {"dev": "arn:same", "prod": "arn:same"}, ["dev", "prod"], payload,
            )
        self.assertEqual(
            {"dev": "Feishu notification configuration error", "prod": "Feishu notification configuration error"},
            results,
        )
        secrets.get_secret_value.assert_not_called()
        request.assert_not_called()

    def test_notification_infrastructure_and_backfill_contract(self):
        with open(os.path.join(ROOT, "infrastructure", "cloudformation.yaml"), encoding="utf-8") as template_file:
            template = template_file.read()
        self.assertIn("FeishuDevSecretArn:", template)
        self.assertIn("FeishuProdSecretArn:", template)
        self.assertIn("FEISHU_DEV_SECRET_ARN: !Ref FeishuDevSecretArn", template)
        self.assertIn("FEISHU_PROD_SECRET_ARN: !Ref FeishuProdSecretArn", template)
        self.assertIn("MonthlyFinalNotificationEnabled:", template)
        self.assertIn("MonthlyFinalNotificationChannel:", template)
        self.assertIn('"notify": ${MonthlyFinalNotificationEnabled}', template)
        self.assertIn('"notification_channel": "${MonthlyFinalNotificationChannel}"', template)
        self.assertNotIn("FEISHU_SECRET_ARN:", template)
        self.assertEqual(
            {
                "month": "2026-07", "report_type": "final", "notify": True,
                "notification_channel": "dev", "backfill": True,
            },
            backfill.build_invocation_payload("2026-07", True),
        )
        self.assertEqual(
            "prod", backfill.build_invocation_payload("2026-07", True, "prod")["notification_channel"],
        )
        with self.assertRaises(ValueError):
            backfill.build_invocation_payload("2026-07", True, "production")

    def test_backfill_range_and_long_timeout(self):
        backfill.validate_range("2026-02", "2026-07", "2026-08", False)
        backfill.validate_range("2026-02", "2026-08", "2026-08", True)
        with self.assertRaises(SystemExit):
            backfill.validate_range("2026-02", "2026-09", "2026-08", True)
        with self.assertRaises(SystemExit):
            backfill.validate_range("2026-02", "2026-08", "2026-08", False)
        config = backfill.lambda_client_config()
        self.assertGreater(config.read_timeout, 900)
        self.assertEqual(10, config.connect_timeout)


if __name__ == "__main__":
    unittest.main()
