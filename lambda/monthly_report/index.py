"""Kiro monthly credits report Lambda.

The module deliberately keeps monthly reporting independent from the existing daily
QuickSight/PDF path. AWS clients are created lazily so helpers can be tested locally.
"""
from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import json
import os
import re
import time
import urllib.request
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Iterable

import boto3
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

TIERS = {
    "Unknown": {"rank": 0, "capacity": 0, "price": 0},
    "Pro": {"rank": 1, "capacity": 1000, "price": 20},
    "Pro+": {"rank": 2, "capacity": 2000, "price": 40},
    "Pro Max": {"rank": 3, "capacity": 5000, "price": 100},
    "Power": {"rank": 4, "capacity": 10000, "price": 200},
}
PARTIAL_START = date(2026, 2, 10)
IDENTITY_MAPPING_KEY = "user-mapping/user_mapping.csv"
SHEETS = ["月度用户明细", "零使用与低使用用户", "订阅类型汇总", "部门汇总", "身份与数据异常", "指标说明"]
ILLEGAL_XML = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")
FORMULA_PREFIXES = ("=", "+", "-", "@")
CSV_ALIASES = {
    "userid": ("userid", "user_id"),
    "user_name": ("user_name", "username", "name"),
    "email": ("email",),
    "subscription_status": ("subscription_status", "status"),
    "subscription_tier": ("subscription_tier", "kiro_plan", "plan"),
    "activation_date": ("activation_date",),
    "plan_source": ("plan_source", "source"),
}
DETAIL_FIELDS = [
    "month", "user_id", "user_name", "email", "department", "subscription_status",
    "plan_source", "tier_history", "month_end_tier", "credits", "overage", "capacity",
    "usage_rate", "color", "capacity_pressure", "estimated_plan_cost", "unused_capacity_value",
    "activation_date", "new_subscription", "first_active", "last_active", "active_days",
    "consecutive_low_months", "consecutive_zero_months", "previous_month_credits", "mom_change",
    "anomaly",
]
DETAIL_NUMERIC_FIELDS = {
    "credits", "overage", "capacity", "usage_rate", "estimated_plan_cost",
    "unused_capacity_value", "active_days", "consecutive_low_months",
    "consecutive_zero_months", "previous_month_credits", "mom_change",
}
HEADERS = [
    "用户ID", "姓名", "邮箱", "部门", "订阅状态", "计划来源", "层级历史", "月末层级",
    "Credits", "容量", "使用率", "预计计划成本(USD)", "未用容量价值(USD)", "超额Credits",
    "激活日期", "新订阅观察", "首次活跃", "最后活跃", "活跃天数", "连续低使用月",
    "连续零使用月", "上月Credits", "环比", "容量压力", "异常",
]


def sanitize(value: Any) -> str:
    """Remove illegal XML/control chars and neutralize spreadsheet formula injection."""
    if value is None:
        return ""
    text = ILLEGAL_XML.sub("", str(value)).replace("\r", " ").replace("\n", " ").strip()
    if text.startswith(FORMULA_PREFIXES):
        text = "'" + text
    return text


def canonical_user_id(value: Any) -> str:
    value = sanitize(value)
    if value.startswith("d-") and "." in value:
        value = value.split(".", 1)[1]
    return value.strip()


def normalize_tier(value: Any) -> str:
    raw = sanitize(value).lower()
    compact = re.sub(r"[^a-z0-9+]", "", raw)
    if not compact:
        return "Unknown"
    if "power" in compact:
        return "Power"
    if "promax" in compact or compact in {"max", "kiropromaxusers"}:
        return "Pro Max"
    if "proplus" in compact or "pro+" in raw or compact in {"plus", "kiroprousers+"}:
        return "Pro+"
    # Exact group names normalize to kiroproplususers / kiroprousers.
    if compact == "kiroproplususers":
        return "Pro+"
    if compact in {"pro", "kiroprousers"} or compact.startswith("pro"):
        return "Pro"
    return "Unknown"


def highest_tier(values: Iterable[Any]) -> str:
    tiers = [normalize_tier(v) for v in values]
    return max(tiers or ["Unknown"], key=lambda t: TIERS[t]["rank"])


def usage_band(credits: Any, capacity: Any) -> tuple[str, str]:
    credits, capacity = decimal(credits), decimal(capacity)
    rate = credits / capacity if capacity > 0 else Decimal("0")
    color = "red" if credits == 0 or rate < Decimal("0.10") else "yellow" if rate < Decimal("0.50") else "green"
    pressure = "exceeded" if capacity > 0 and rate >= 1 else "capacity_pressure" if capacity > 0 and rate >= Decimal("0.90") else ""
    return color, pressure


def decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def month_add(month: str, delta: int) -> str:
    dt = datetime.strptime(month, "%Y-%m")
    idx = dt.year * 12 + dt.month - 1 + delta
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}"


def month_bounds(month: str) -> tuple[date, date]:
    start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    return start, (start.replace(day=28) + timedelta(days=4)).replace(day=1)


def _event_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value)
    normalized = sanitize(value).lower()
    if normalized in {"true", "1", "yes", "on"}:
        return True
    if normalized in {"false", "0", "no", "off"}:
        return False
    raise ValueError("notify must be a boolean")


def select_notification_channels(event: dict[str, Any]) -> list[str]:
    """Select explicit delivery channels; omission is always safe development."""
    channel = sanitize(event.get("notification_channel") or "dev").lower()
    if channel == "both":
        return ["dev", "prod"]
    if channel not in {"dev", "prod"}:
        raise ValueError("notification_channel must be dev, prod, or both")
    return [channel]


def select_report_month(event: dict[str, Any], now: datetime | None = None) -> dict[str, Any]:
    """Select target month. Scheduled events always use event.time, never Lambda wall clock."""
    if event.get("month"):
        month = str(event["month"])
        if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month):
            raise ValueError("month must be YYYY-MM")
        kind = sanitize(event.get("report_type") or "final").lower()
        if kind not in {"provisional", "final"}:
            raise ValueError("report_type must be provisional or final")
        return {
            "month": month, "report_type": kind, "scheduled": False,
            "notify": _event_bool(event.get("notify"), False),
        }
    raw_time = event.get("time")
    if raw_time:
        event_time = datetime.fromisoformat(str(raw_time).replace("Z", "+00:00"))
    else:
        event_time = now or datetime.now(timezone.utc)
    current = f"{event_time.year:04d}-{event_time.month:02d}"
    kind = sanitize(event.get("report_type") or ("provisional" if event_time.day == 1 else "final")).lower()
    if kind not in {"provisional", "final"}:
        raise ValueError("report_type must be provisional or final")
    return {
        "month": month_add(current, -1), "report_type": kind, "scheduled": True,
        "notify": _event_bool(event.get("notify"), False),
    }


def report_status(month: str) -> tuple[str, str]:
    if month == "2026-02":
        return "PARTIAL", "2026-02 数据自 2026-02-10 起可用，月度数据不完整。"
    return "COMPLETE", ""


def monthly_keys(prefix: str, month: str, report_type: str) -> dict[str, str]:
    year, mon = month.split("-")
    base = f"{prefix.strip('/')}/{year}/{mon}"
    return {
        "workbook": f"{base}/kiro-credits-{month}-{report_type}.xlsx",
        "alias": f"{base}/kiro-credits-{month}.xlsx",
        "detail_csv": f"{base}/kiro-credits-{month}-detail.csv",
        "subscription_csv": f"{base}/kiro-credits-{month}-subscriptions.csv",
        "annual": f"{prefix.strip('/')}/{year}/kiro-credits-{year}.xlsx",
    }


def _canonical_csv_row(row: dict[str, Any]) -> dict[str, str]:
    lowered = {sanitize(k).lower(): sanitize(v) for k, v in row.items() if k is not None}
    result = {}
    for canonical, aliases in CSV_ALIASES.items():
        result[canonical] = next((lowered[a] for a in aliases if a in lowered), "")
    result["userid"] = canonical_user_id(result["userid"])
    result["subscription_tier"] = normalize_tier(result["subscription_tier"])
    result["subscription_status"] = result["subscription_status"] or "ACTIVE"
    result["plan_source"] = result["plan_source"] or "S3 CSV"
    return result


def parse_subscription_csv(data: bytes) -> dict[str, dict[str, str]]:
    """Parse strict UTF-8-sig CSV. Header-only/valid empty CSV is authoritative."""
    text = data.decode("utf-8-sig", errors="strict")
    reader = csv.DictReader(io.StringIO(text, newline=""))
    if reader.fieldnames is None:
        raise ValueError("subscription CSV requires a header")
    known = {a for aliases in CSV_ALIASES.values() for a in aliases}
    headers = {sanitize(h).lower() for h in reader.fieldnames if h}
    if not headers.intersection(CSV_ALIASES["userid"]):
        raise ValueError("subscription CSV requires userid or user_id")
    if not headers.issubset(known):
        # Extra fields are ignored intentionally, but malformed canonical headers should not silently pass.
        pass
    roster: dict[str, dict[str, str]] = {}
    for raw in reader:
        row = _canonical_csv_row(raw)
        uid = row["userid"]
        if not uid:
            continue
        prior = roster.get(uid)
        if prior is None:
            roster[uid] = row
            continue
        selected, other = (row, prior) if TIERS[row["subscription_tier"]]["rank"] >= TIERS[prior["subscription_tier"]]["rank"] else (prior, row)
        selected = dict(selected)
        for field in ("user_name", "email"):
            value = sanitize(selected.get(field))
            fallback = sanitize(other.get(field))
            if (not value or (field == "user_name" and value == uid)) and fallback and not (field == "user_name" and fallback == uid):
                selected[field] = fallback
        roster[uid] = selected
    return roster


def load_csv_roster(s3: Any, bucket: str, key: str) -> tuple[dict[str, dict[str, str]] | None, bool]:
    """Return (roster, authoritative). Only an absent exact key permits API fallback."""
    if not key:
        return None, False
    try:
        body = s3.get_object(Bucket=bucket, Key=key)["Body"].read()
    except Exception as exc:
        code = getattr(exc, "response", {}).get("Error", {}).get("Code", "")
        if code in {"NoSuchKey", "404", "NotFound"}:
            return None, False
        raise
    return parse_subscription_csv(body), True


def parse_identity_mapping_csv(data: bytes) -> dict[str, dict[str, str]]:
    """Parse an identity-only mapping; subscription fields are intentionally impossible to return."""
    reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig", errors="strict"), newline=""))
    if reader.fieldnames is None:
        raise ValueError("identity mapping CSV requires a header")
    result: dict[str, dict[str, str]] = {}
    for raw in reader:
        lowered = {sanitize(k).lower(): sanitize(v) for k, v in raw.items() if k is not None}
        uid = canonical_user_id(lowered.get("userid") or lowered.get("user_id"))
        if not uid:
            continue
        name = sanitize(lowered.get("user_name") or lowered.get("username") or lowered.get("name"))
        email = sanitize(lowered.get("email"))
        identity = result.setdefault(uid, {"user_name": "", "email": ""})
        if name and name != uid:
            identity["user_name"] = name
        if email:
            identity["email"] = email
    return result


def load_identity_mapping(s3: Any, bucket: str, key: str = IDENTITY_MAPPING_KEY) -> dict[str, dict[str, str]]:
    try:
        body = s3.get_object(Bucket=bucket, Key=key)["Body"].read()
    except Exception as exc:
        code = getattr(exc, "response", {}).get("Error", {}).get("Code", "")
        if code in {"NoSuchKey", "404", "NotFound"}:
            return {}
        raise
    return parse_identity_mapping_csv(body)


def identity_projection(rows: dict[str, dict[str, Any]]) -> dict[str, dict[str, str]]:
    """Project only name/email so current subscription facts cannot leak into history."""
    result: dict[str, dict[str, str]] = {}
    for raw_uid, row in rows.items():
        uid = canonical_user_id(raw_uid or row.get("userid") or row.get("user_id"))
        if not uid:
            continue
        name = sanitize(row.get("user_name") or row.get("username"))
        result[uid] = {"user_name": name if name != uid else "", "email": sanitize(row.get("email"))}
    return result


def identities_from_history(history: dict[str, list[dict[str, str]]]) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for uid, observations in history.items():
        identity = {"user_name": "", "email": ""}
        for row in observations:
            name = sanitize(row.get("user_name"))
            if name and name != uid:
                identity["user_name"] = name
            email = sanitize(row.get("email"))
            if email:
                identity["email"] = email
        result[uid] = identity
    return result


def merge_identity_maps(*maps: dict[str, dict[str, str]]) -> dict[str, dict[str, str]]:
    """Merge field-by-field; later non-empty identity observations take precedence."""
    result: dict[str, dict[str, str]] = {}
    for mapping in maps:
        for raw_uid, identity in mapping.items():
            uid = canonical_user_id(raw_uid)
            if not uid:
                continue
            target = result.setdefault(uid, {"user_name": "", "email": ""})
            for field in ("user_name", "email"):
                value = sanitize(identity.get(field))
                if value and not (field == "user_name" and value == uid):
                    target[field] = value
    return result


def _paginate(client: Any, method: str, result_key: str, **kwargs: Any) -> Iterable[dict[str, Any]]:
    token = None
    while True:
        request = dict(kwargs)
        if token:
            request["NextToken"] = token
        response = getattr(client, method)(**request)
        yield from response.get(result_key, [])
        token = response.get("NextToken")
        if not token:
            break


def _describe_user(identitystore: Any, identity_store_id: str, uid: str) -> dict[str, str]:
    try:
        user = identitystore.describe_user(IdentityStoreId=identity_store_id, UserId=uid)
    except Exception:
        return {"user_name": uid, "email": ""}
    emails = user.get("Emails", [])
    primary = next((e.get("Value", "") for e in emails if e.get("Primary")), "") or (emails[0].get("Value", "") if emails else "")
    return {"user_name": sanitize(user.get("DisplayName") or user.get("UserName") or uid), "email": sanitize(primary)}


def resolve_application_roster(sso: Any, identitystore: Any, application_arn: str, identity_store_id: str,
                               latest_usage_tiers: dict[str, str]) -> dict[str, dict[str, str]]:
    if not application_arn:
        raise ValueError("KIRO_APPLICATION_ARN is required when subscription CSV is absent")
    direct_users: set[str] = set()
    group_tiers: dict[str, list[str]] = defaultdict(list)
    for assignment in _paginate(sso, "list_application_assignments", "ApplicationAssignments", ApplicationArn=application_arn):
        principal_id = canonical_user_id(assignment.get("PrincipalId"))
        if assignment.get("PrincipalType") == "USER":
            direct_users.add(principal_id)
        elif assignment.get("PrincipalType") == "GROUP":
            group = identitystore.describe_group(IdentityStoreId=identity_store_id, GroupId=principal_id)
            tier = normalize_tier(group.get("DisplayName"))
            for membership in _paginate(identitystore, "list_group_memberships", "GroupMemberships",
                                         IdentityStoreId=identity_store_id, GroupId=principal_id):
                uid = canonical_user_id(membership.get("MemberId", {}).get("UserId"))
                if uid:
                    group_tiers[uid].append(tier)
    users = direct_users | set(group_tiers)
    roster = {}
    for uid in sorted(users):
        info = _describe_user(identitystore, identity_store_id, uid)
        tiers = group_tiers.get(uid) or [latest_usage_tiers.get(uid, "Unknown")]
        roster[uid] = {
            "userid": uid, **info, "subscription_status": "ACTIVE",
            "subscription_tier": highest_tier(tiers), "activation_date": "",
            "plan_source": "IIC group" if uid in group_tiers else "IIC direct assignment",
        }
    return roster


def _tier_sql() -> str:
    compact = "regexp_replace(upper(COALESCE(subscription_tier, '')), '[^A-Z0-9]', '')"
    return (f"CASE WHEN {compact} LIKE '%POWER%' THEN 'Power' "
            f"WHEN {compact} LIKE '%PROMAX%' THEN 'Pro Max' "
            f"WHEN {compact} LIKE '%PROPLUS%' THEN 'Pro+' "
            f"WHEN {compact} LIKE '%PRO%' THEN 'Pro' ELSE 'Unknown' END")


def _tier_rank_sql(label: str = "tier_label") -> str:
    return f"CASE {label} WHEN 'Power' THEN 4 WHEN 'Pro Max' THEN 3 WHEN 'Pro+' THEN 2 WHEN 'Pro' THEN 1 ELSE 0 END"

def build_usage_sql(database: str, month: str) -> str:
    start, end = month_bounds(month)
    tier_sql, rank_sql = _tier_sql(), _tier_rank_sql()
    return f"""
WITH typed AS (
 SELECT CASE WHEN strpos(userid, '.') > 0 THEN split_part(userid, '.', 2) ELSE userid END AS user_id,
        TRY(date_parse(date, '%Y-%m-%d')) AS usage_date,
        TRY_CAST(credits_used AS DECIMAL(18,4)) AS credits,
        TRY_CAST(overage_credits_used AS DECIMAL(18,4)) AS overage,
        {tier_sql} AS tier_label
 FROM \"{database}\".\"user_report\"
 WHERE TRY(date_parse(date, '%Y-%m-%d')) >= TIMESTAMP '{start.isoformat()} 00:00:00'
   AND TRY(date_parse(date, '%Y-%m-%d')) < TIMESTAMP '{end.isoformat()} 00:00:00'
), ranked AS (
 SELECT *, {rank_sql} AS tier_rank
 FROM typed WHERE user_id IS NOT NULL AND user_id <> '' AND usage_date IS NOT NULL
), by_day AS (
 SELECT user_id, usage_date, SUM(COALESCE(credits, 0)) credits,
        SUM(COALESCE(overage, 0)) overage,
        CASE MAX(tier_rank) WHEN 4 THEN 'Power' WHEN 3 THEN 'Pro Max'
             WHEN 2 THEN 'Pro+' WHEN 1 THEN 'Pro' ELSE 'Unknown' END AS day_tier
 FROM ranked GROUP BY user_id, usage_date
), history AS (
 SELECT user_id, array_join(array_sort(array_distinct(array_agg(tier_label))), ' -> ') tier_history
 FROM ranked GROUP BY user_id
)
SELECT d.user_id, CAST(SUM(d.credits) AS VARCHAR) credits, CAST(SUM(d.overage) AS VARCHAR) overage,
       CAST(MIN(d.usage_date) AS VARCHAR) first_active, CAST(MAX(d.usage_date) AS VARCHAR) last_active,
       CAST(COUNT(DISTINCT d.usage_date) AS VARCHAR) active_days, h.tier_history,
       max_by(d.day_tier, d.usage_date) latest_tier
FROM by_day d JOIN history h ON d.user_id = h.user_id
GROUP BY d.user_id, h.tier_history ORDER BY d.user_id
""".strip()


def build_latest_tiers_sql(database: str) -> str:
    tier_sql, rank_sql = _tier_sql(), _tier_rank_sql()
    return f"""
WITH typed AS (
 SELECT CASE WHEN strpos(userid, '.') > 0 THEN split_part(userid, '.', 2) ELSE userid END AS user_id,
        TRY(date_parse(date, '%Y-%m-%d')) AS usage_date,
        {tier_sql} AS tier_label
 FROM \"{database}\".\"user_report\"
), ranked AS (
 SELECT *, {rank_sql} AS tier_rank
 FROM typed WHERE user_id IS NOT NULL AND user_id <> '' AND usage_date IS NOT NULL
), by_day AS (
 SELECT user_id, usage_date,
        CASE MAX(tier_rank) WHEN 4 THEN 'Power' WHEN 3 THEN 'Pro Max'
             WHEN 2 THEN 'Pro+' WHEN 1 THEN 'Pro' ELSE 'Unknown' END AS day_tier
 FROM ranked GROUP BY user_id, usage_date
)
SELECT user_id, max_by(day_tier, usage_date) AS latest_tier
FROM by_day GROUP BY user_id ORDER BY user_id
""".strip()


def run_athena_query(athena: Any, sql: str, workgroup: str, timeout_seconds: int = 720) -> list[dict[str, str]]:
    qid = athena.start_query_execution(QueryString=sql, WorkGroup=workgroup)["QueryExecutionId"]
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        status = athena.get_query_execution(QueryExecutionId=qid)["QueryExecution"]["Status"]
        state = status["State"]
        if state == "SUCCEEDED":
            break
        if state in {"FAILED", "CANCELLED"}:
            raise RuntimeError(f"Athena query {state}: {status.get('StateChangeReason', '')}")
        time.sleep(1)
    else:
        try:
            athena.stop_query_execution(QueryExecutionId=qid)
        finally:
            raise TimeoutError(f"Athena query timed out after {timeout_seconds}s")
    rows: list[list[str]] = []
    token = None
    while True:
        kwargs = {"QueryExecutionId": qid}
        if token:
            kwargs["NextToken"] = token
        page = athena.get_query_results(**kwargs)
        for row in page.get("ResultSet", {}).get("Rows", []):
            rows.append([cell.get("VarCharValue", "") for cell in row.get("Data", [])])
        token = page.get("NextToken")
        if not token:
            break
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row + [""] * (len(headers) - len(row)))) for row in rows[1:]]


def usage_rows_by_id(rows: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Merge by canonical ID only; duplicate display names never merge."""
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        uid = canonical_user_id(row.get("user_id") or row.get("userid"))
        if not uid:
            continue
        item = result.setdefault(uid, {"user_id": uid, "credits": Decimal(0), "overage": Decimal(0), "tier_history": [], "active_days": 0})
        item["credits"] += decimal(row.get("credits"))
        item["overage"] += decimal(row.get("overage"))
        item["active_days"] += int(decimal(row.get("active_days")))
        item["first_active"] = min(filter(None, [item.get("first_active"), sanitize(row.get("first_active"))]), default="")
        item["last_active"] = max(filter(None, [item.get("last_active"), sanitize(row.get("last_active"))]), default="")
        history = [normalize_tier(t) for t in re.split(r"\s*(?:->|→)\s*", sanitize(row.get("tier_history"))) if t]
        item["tier_history"] = sorted(set(item["tier_history"] + history), key=lambda t: TIERS[t]["rank"])
        item["latest_tier"] = normalize_tier(row.get("latest_tier"))
    return result


def latest_usage_tiers(usage: dict[str, dict[str, Any]]) -> dict[str, str]:
    return {uid: item.get("latest_tier", "Unknown") for uid, item in usage.items()}


def should_use_current_roster(month: str, current_month: str) -> bool:
    return month >= month_add(current_month, -1)


def _department(email: str) -> str:
    return email.rsplit("@", 1)[1].lower() if "@" in email else "未识别"


def should_write_subscription_snapshot(include_roster: bool) -> bool:
    """Only persist a roster when it is valid for the reported recent month."""
    return include_roster


def merge_report_rows(month: str, usage: dict[str, dict[str, Any]], roster: dict[str, dict[str, Any]],
                      include_roster: bool, previous: dict[str, dict[str, str]] | None = None,
                      history: dict[str, list[dict[str, str]]] | None = None,
                      identity_map: dict[str, dict[str, str]] | None = None) -> list[dict[str, Any]]:
    previous, history, identity_map = previous or {}, history or {}, identity_map or {}
    ids = set(usage) | (set(roster) if include_roster else set())
    rows = []
    for uid in sorted(ids):
        u = usage.get(uid, {})
        sub = roster.get(uid, {}) if include_roster else {}
        observed_tier = normalize_tier(u.get("latest_tier"))
        tier = observed_tier if observed_tier != "Unknown" else normalize_tier(sub.get("subscription_tier"))
        meta = TIERS[tier]
        credits = decimal(u.get("credits"))
        capacity = Decimal(meta["capacity"])
        rate = credits / capacity if capacity else Decimal(0)
        color, pressure = usage_band(credits, capacity)
        prev_credits = decimal(previous.get(uid, {}).get("credits"))
        mom = (credits - prev_credits) / prev_credits if prev_credits else None
        past = history.get(uid, [])
        low_count = zero_count = 0
        for old in reversed(past):
            if sanitize(old.get("color")) in {"red", "yellow"}:
                low_count += 1
            else:
                break
        for old in reversed(past):
            if decimal(old.get("credits")) == 0:
                zero_count += 1
            else:
                break
        if color in {"red", "yellow"}: low_count += 1
        if credits == 0: zero_count += 1
        activation = sanitize(sub.get("activation_date"))
        new_flag = "是" if activation.startswith(month) else ""
        anomaly = []
        if tier == "Unknown": anomaly.append("未知订阅层级")
        if not include_roster or uid not in roster: anomaly.append("仅有使用记录/订阅快照缺失")
        if include_roster and uid not in usage: anomaly.append("订阅用户无使用记录")
        identity = identity_map.get(uid, {})
        user_name = sanitize(identity.get("user_name") or sub.get("user_name") or uid)
        email = sanitize(identity.get("email") or sub.get("email"))
        rows.append({
            "month": month, "user_id": uid, "user_name": user_name, "email": email,
            "department": _department(email), "subscription_status": sanitize(sub.get("subscription_status") or ("ACTIVE" if sub else "UNKNOWN")),
            "plan_source": sanitize(sub.get("plan_source") or "usage report"),
            "tier_history": " → ".join(u.get("tier_history") or ([tier] if tier != "Unknown" else [])),
            "month_end_tier": tier, "credits": credits, "overage": decimal(u.get("overage")),
            "capacity": int(capacity), "usage_rate": rate, "color": color, "capacity_pressure": pressure,
            "estimated_plan_cost": meta["price"],
            "unused_capacity_value": Decimal(meta["price"]) * max(Decimal(0), Decimal(1) - rate) if capacity else Decimal(0),
            "activation_date": activation, "new_subscription": new_flag,
            "first_active": sanitize(u.get("first_active")), "last_active": sanitize(u.get("last_active")),
            "active_days": int(u.get("active_days", 0)), "consecutive_low_months": low_count,
            "consecutive_zero_months": zero_count, "previous_month_credits": prev_credits,
            "mom_change": mom, "anomaly": "; ".join(anomaly),
        })
    rows.sort(key=lambda row: (
        decimal(row["usage_rate"]), decimal(row["credits"]),
        str(row["user_name"]).casefold(), str(row["user_id"]),
    ))
    return rows


def _csv_field_value(field: str, value: Any) -> str:
    if value is None:
        return ""
    if field in DETAIL_NUMERIC_FIELDS:
        return ILLEGAL_XML.sub("", str(value)).replace("\r", " ").replace("\n", " ").strip()
    return sanitize(value)


def rows_to_csv(rows: list[dict[str, Any]], fields: list[str] = DETAIL_FIELDS) -> bytes:
    out = io.StringIO(newline="")
    writer = csv.DictWriter(out, fieldnames=fields, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _csv_field_value(key, row.get(key)) for key in fields})
    return out.getvalue().encode("utf-8-sig")


def csv_to_rows(data: bytes) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig", errors="strict")))
    return [{key: _csv_field_value(key, value) for key, value in row.items()} for row in reader]


def subscription_snapshot_csv(roster: dict[str, dict[str, Any]]) -> bytes:
    fields = ["userid", "user_name", "email", "subscription_status", "subscription_tier", "activation_date", "plan_source"]
    return rows_to_csv([roster[k] for k in sorted(roster)], fields)


NAVY = "17365D"; BLUE = "5B9BD5"; RED = "F4CCCC"; YELLOW = "FFF2CC"; GREEN = "D9EAD3"; WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E2F3")


def _set_calc_mode(wb: Workbook) -> None:
    calc = getattr(wb, "calculation", None) or getattr(wb, "calculation_properties", None)
    if calc:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
        calc.calcMode = "auto"


def _style_sheet(ws: Any, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    # Excel Tables own their AutoFilter. A duplicate worksheet AutoFilter over the
    # same range can make Microsoft Excel repair/delete workbook view records.
    ws.auto_filter.ref = None
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="center")
    for column in range(1, ws.max_column + 1):
        values = [len(str(ws.cell(r, column).value or "")) for r in range(1, min(ws.max_row, 200) + 1)]
        ws.column_dimensions[get_column_letter(column)].width = min(45, max(10, max(values, default=10) + 2))
    if ws.max_row >= 2 and ws.max_column >= 1:
        name = hashlib.sha1(ws.title.encode("utf-8")).hexdigest()[:12]
        table = Table(displayName=f"T{name}", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)


def _append_detail(ws: Any, rows: list[dict[str, Any]], formulas: bool = True) -> None:
    ws.append(HEADERS)
    for idx, row in enumerate(rows, 2):
        ws.append([
            sanitize(row["user_id"]), sanitize(row["user_name"]), sanitize(row["email"]), sanitize(row["department"]),
            sanitize(row["subscription_status"]), sanitize(row["plan_source"]), sanitize(row["tier_history"]), row["month_end_tier"],
            float(decimal(row["credits"])), int(row["capacity"]), None if formulas else float(decimal(row["usage_rate"])),
            float(decimal(row["estimated_plan_cost"])), None if formulas else float(decimal(row["unused_capacity_value"])),
            float(decimal(row["overage"])), sanitize(row["activation_date"]), sanitize(row["new_subscription"]),
            sanitize(row["first_active"]), sanitize(row["last_active"]), int(decimal(row["active_days"])),
            int(decimal(row["consecutive_low_months"])), int(decimal(row["consecutive_zero_months"])),
            float(decimal(row["previous_month_credits"])), None if row.get("mom_change") is None else float(decimal(row["mom_change"])),
            sanitize(row["capacity_pressure"]), sanitize(row["anomaly"]),
        ])
        if formulas:
            ws.cell(idx, 11, f"=IFERROR(IF(J{idx}>0,I{idx}/J{idx},0),0)")
            ws.cell(idx, 13, f"=IFERROR(IF(J{idx}>0,L{idx}*MAX(0,1-K{idx}),0),0)")
    _style_sheet(ws)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 11).number_format = "0.0%"
        ws.cell(row, 13).number_format = "$0.00"
        ws.cell(row, 23).number_format = "0.0%;-0.0%;-"
        color = rows[row - 2].get("color")
        ws.cell(row, 11).fill = PatternFill("solid", fgColor={"red": RED, "yellow": YELLOW, "green": GREEN}.get(color, WHITE))
    if ws.max_row >= 2:
        ws.conditional_formatting.add(f"K2:K{ws.max_row}", CellIsRule(operator="lessThan", formula=["0.1"], fill=PatternFill("solid", fgColor=RED)))
        ws.conditional_formatting.add(f"K2:K{ws.max_row}", CellIsRule(operator="between", formula=["0.1", "0.499999"], fill=PatternFill("solid", fgColor=YELLOW)))
        ws.conditional_formatting.add(f"K2:K{ws.max_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0.5"], fill=PatternFill("solid", fgColor=GREEN)))


def build_monthly_workbook(rows: list[dict[str, Any]], month: str, status: str, warning: str = "") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    detail = wb.create_sheet(SHEETS[0]); _append_detail(detail, rows)
    low = wb.create_sheet(SHEETS[1]); _append_detail(low, [r for r in rows if r["color"] in {"red", "yellow"}])
    tier = wb.create_sheet(SHEETS[2]); tier.append(["订阅层级", "用户数", "Credits", "容量", "使用率", "计划成本(USD)"])
    for name in ["Pro", "Pro+", "Pro Max", "Power", "Unknown"]:
        subset = [r for r in rows if r["month_end_tier"] == name]
        credits = sum((decimal(r["credits"]) for r in subset), Decimal(0)); capacity = sum(int(r["capacity"]) for r in subset)
        tier.append([name, len(subset), float(credits), capacity, float(credits / capacity) if capacity else 0, sum(decimal(r["estimated_plan_cost"]) for r in subset)])
    _style_sheet(tier)
    for cell in tier["E"][1:]: cell.number_format = "0.0%"
    dept = wb.create_sheet(SHEETS[3]); dept.append(["部门/邮箱域", "用户数", "Credits", "零使用用户", "低使用用户", "计划成本(USD)"])
    departments = sorted({r["department"] for r in rows})
    for name in departments:
        subset = [r for r in rows if r["department"] == name]
        dept.append([sanitize(name), len(subset), float(sum((decimal(r["credits"]) for r in subset), Decimal(0))), sum(decimal(r["credits"]) == 0 for r in subset), sum(r["color"] in {"red", "yellow"} for r in subset), float(sum((decimal(r["estimated_plan_cost"]) for r in subset), Decimal(0)))])
    _style_sheet(dept)
    anomalies = wb.create_sheet(SHEETS[4]); anomalies.append(["用户ID", "姓名", "异常"])
    for row in rows:
        if row["anomaly"]: anomalies.append([sanitize(row["user_id"]), sanitize(row["user_name"]), sanitize(row["anomaly"])])
    _style_sheet(anomalies)
    notes = wb.create_sheet(SHEETS[5]); notes.append(["项目", "说明"])
    notes_rows = [
        ("报告月份", month), ("数据状态", status), ("数据警告", warning or "无"),
        ("Credits范围", "仅查询 Athena user_report，按规范化用户ID汇总，日期使用半开区间。"),
        ("红/黄/绿", "红：零使用或<10%；黄：10%-<50%；绿：>=50%。>=90%容量压力，>=100%超额。"),
        ("官方套餐", "Pro $20/1000；Pro+ $40/2000；Pro Max $100/5000；Power $200/10000；Unknown $0/0。"),
        ("历史订阅", "历史回填若早于上个自然月，不使用当前订阅快照补零，以免误称为历史订阅。"),
        ("身份退避", "历史姓名/邮箱可使用当前或最后已知 ID 映射，仅用于身份显示，不代表报告月订阅关系。"),
        ("新订阅观察", "激活日期位于报告月，或连续历史快照中首次观察到该订阅。"),
    ]
    for item in notes_rows: notes.append([sanitize(item[0]), sanitize(item[1])])
    _style_sheet(notes)
    notes.column_dimensions["B"].width = 100
    for ws in wb.worksheets: ws.sheet_view.showGridLines = False
    _set_calc_mode(wb)
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def build_annual_workbook(monthly_rows: dict[str, list[dict[str, Any]]], year: str) -> bytes:
    """Rebuild annual output solely from complete stored monthly snapshots."""
    wb = Workbook(); summary = wb.active; summary.title = "年度汇总"
    summary.append(["月份", "用户数", "零使用", "红", "黄", "绿", "Credits", "预计计划成本(USD)"])
    for month in sorted(monthly_rows):
        rows = monthly_rows[month]
        summary.append([month, len(rows), sum(decimal(r.get("credits")) == 0 for r in rows), sum(r.get("color") == "red" for r in rows), sum(r.get("color") == "yellow" for r in rows), sum(r.get("color") == "green" for r in rows), float(sum((decimal(r.get("credits")) for r in rows), Decimal(0))), float(sum((decimal(r.get("estimated_plan_cost")) for r in rows), Decimal(0)))])
        ws = wb.create_sheet(month)
        normalized = []
        for raw in rows:
            item = {field: raw.get(field, "") for field in DETAIL_FIELDS}
            item.update({"credits": decimal(item["credits"]), "overage": decimal(item["overage"]), "capacity": int(decimal(item["capacity"])), "estimated_plan_cost": decimal(item["estimated_plan_cost"]), "unused_capacity_value": decimal(item["unused_capacity_value"]), "active_days": int(decimal(item["active_days"])), "consecutive_low_months": int(decimal(item["consecutive_low_months"])), "consecutive_zero_months": int(decimal(item["consecutive_zero_months"])), "previous_month_credits": decimal(item["previous_month_credits"]), "mom_change": decimal(item["mom_change"]) if item["mom_change"] not in {"", None} else None})
            normalized.append(item)
        _append_detail(ws, normalized, formulas=False)
    _style_sheet(summary)
    _set_calc_mode(wb)
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def list_monthly_snapshots(s3: Any, bucket: str, prefix: str, year: str) -> dict[str, list[dict[str, str]]]:
    snapshots = {}
    token = None
    base = f"{prefix.strip('/')}/{year}/"
    while True:
        kwargs = {"Bucket": bucket, "Prefix": base}
        if token: kwargs["ContinuationToken"] = token
        page = s3.list_objects_v2(**kwargs)
        for obj in page.get("Contents", []):
            match = re.search(r"kiro-credits-(\d{4}-\d{2})-detail\.csv$", obj["Key"])
            if match:
                snapshots[match.group(1)] = csv_to_rows(s3.get_object(Bucket=bucket, Key=obj["Key"])["Body"].read())
        token = page.get("NextContinuationToken")
        if not token: break
    return snapshots


def load_history(s3: Any, bucket: str, prefix: str, month: str) -> tuple[dict[str, dict[str, str]], dict[str, list[dict[str, str]]]]:
    year = int(month[:4])
    all_rows = {}
    all_rows.update(list_monthly_snapshots(s3, bucket, prefix, str(year - 1)))
    all_rows.update(list_monthly_snapshots(s3, bucket, prefix, str(year)))
    prior_month = month_add(month, -1)
    previous = {canonical_user_id(r.get("user_id")): r for r in all_rows.get(prior_month, [])}
    history: dict[str, list[dict[str, str]]] = defaultdict(list)
    for old_month in sorted(m for m in all_rows if m < month):
        for row in all_rows[old_month]:
            history[canonical_user_id(row.get("user_id"))].append(row)
    return previous, history


def public_url(bucket: str, region: str, key: str) -> str:
    return f"http://{bucket}.s3-website-{region}.amazonaws.com/{key}"


def feishu_signature(secret: str, timestamp: str) -> str:
    key = f"{timestamp}\n{secret}".encode("utf-8")
    return base64.b64encode(hmac.new(key, digestmod=hashlib.sha256).digest()).decode("ascii")


def format_percent(numerator: Any, denominator: Any) -> str:
    total = decimal(denominator)
    if total <= 0:
        return "0%"
    value = (decimal(numerator) * Decimal("100") / total).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return f"{value:.0f}%"


def _format_usage_rate(value: Any) -> str:
    rate = (decimal(value) * Decimal("100")).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    return f"{rate:.1f}%"


def _format_credits(value: Any) -> str:
    amount = decimal(value)
    if amount == amount.to_integral_value():
        return f"{int(amount):,}"
    return f"{amount:,.2f}".rstrip("0").rstrip(".")


def _md_escape(value: Any) -> str:
    text = sanitize(value).replace("\\", "\\\\")
    return re.sub(r"([`*_\[\]])", r"\\\1", text)


def _identity_label(row: dict[str, Any], duplicate_names: set[str]) -> str:
    uid = canonical_user_id(row.get("user_id"))
    name = sanitize(row.get("user_name"))
    if not name or name == uid:
        suffix = uid[-8:] if uid else "未知ID"
        return f"未识别用户（{suffix}）"
    if name.casefold() in duplicate_names:
        qualifier = sanitize(row.get("email")) or (uid[-8:] if uid else "未知ID")
        return f"{name}（{qualifier}）"
    return name


def build_feishu_card(month: str, rows: list[dict[str, Any]], url: str,
                      status: str = "COMPLETE", warning: str = "") -> dict[str, Any]:
    """Build a Card JSON 2.0 report with responsive KPI columns and risk panels."""
    total_users = len(rows)
    zero_rows = [row for row in rows if decimal(row.get("credits")) == 0]
    low_rows = [row for row in rows if decimal(row.get("credits")) > 0 and row.get("color") == "red"]
    yellow_count = sum(row.get("color") == "yellow" for row in rows)
    green_count = sum(row.get("color") == "green" for row in rows)
    total_cost = sum((decimal(row.get("estimated_plan_cost")) for row in rows), Decimal(0))
    zero_cost = sum((decimal(row.get("estimated_plan_cost")) for row in zero_rows), Decimal(0))
    low_idle_value = sum((decimal(row.get("unused_capacity_value")) for row in low_rows), Decimal(0))

    name_counts: dict[str, int] = defaultdict(int)
    for row in rows:
        name = sanitize(row.get("user_name"))
        uid = canonical_user_id(row.get("user_id"))
        if name and name != uid:
            name_counts[name.casefold()] += 1
    duplicate_names = {name for name, count in name_counts.items() if count > 1}

    def label(row: dict[str, Any]) -> str:
        return _identity_label(row, duplicate_names)

    zero_rows.sort(key=lambda row: (
        -int(decimal(row.get("consecutive_zero_months"))),
        -decimal(row.get("estimated_plan_cost")),
        label(row).casefold(), canonical_user_id(row.get("user_id")),
    ))
    low_rows.sort(key=lambda row: (
        decimal(row.get("usage_rate")),
        -int(decimal(row.get("consecutive_low_months"))),
        -decimal(row.get("estimated_plan_cost")),
        label(row).casefold(), canonical_user_id(row.get("user_id")),
    ))

    def metric_column(title: str, value: str, detail: str) -> dict[str, Any]:
        return {
            "tag": "column", "width": "weighted", "weight": 1,
            "background_style": "neutral_bg", "padding": "10px 8px 10px 8px",
            "vertical_align": "center",
            "elements": [{
                "tag": "markdown", "text_align": "center", "text_size": "normal",
                "content": f"{title}\n**{value}**\n{detail}",
            }],
        }

    def metric_row(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
        return {
            "tag": "column_set", "flex_mode": "bisect",
            "horizontal_spacing": "8px", "horizontal_align": "left",
            "columns": [left, right],
        }

    def risk_row(row: dict[str, Any], kind: str) -> dict[str, Any]:
        name = _md_escape(label(row))
        tier = _md_escape(row.get("month_end_tier") or "Unknown")
        cost = decimal(row.get("estimated_plan_cost"))
        if kind == "zero":
            left = f"**{name}**\n连续零使用 **{int(decimal(row.get('consecutive_zero_months')))}个月**"
            right = f"{tier}\n**${cost:.2f}**"
        else:
            left = (
                f"**{name}**\n{_format_credits(row.get('credits'))} / "
                f"{int(decimal(row.get('capacity'))):,} Credits · {_format_usage_rate(row.get('usage_rate'))}"
            )
            right = (
                f"{tier} · **${cost:.2f}**\n"
                f"连续低用 **{int(decimal(row.get('consecutive_low_months')))}个月**"
            )
        return {
            "tag": "column_set", "flex_mode": "none",
            "horizontal_spacing": "8px", "background_style": "default",
            "columns": [
                {
                    "tag": "column", "width": "weighted", "weight": 3,
                    "vertical_align": "center", "padding": "6px 8px 6px 8px",
                    "elements": [{"tag": "markdown", "content": left, "text_size": "normal"}],
                },
                {
                    "tag": "column", "width": "weighted", "weight": 2,
                    "vertical_align": "center", "padding": "6px 8px 6px 8px",
                    "elements": [{
                        "tag": "markdown", "content": right,
                        "text_align": "right", "text_size": "normal",
                    }],
                },
            ],
        }

    risk_count = len(zero_rows) + len(low_rows)

    def panel_rows(items: list[dict[str, Any]], kind: str) -> list[dict[str, Any]]:
        if risk_count > 24:
            compact_rows = []
            for row in items:
                name = _md_escape(label(row))
                tier = _md_escape(row.get("month_end_tier") or "Unknown")
                cost = decimal(row.get("estimated_plan_cost"))
                if kind == "zero":
                    compact_rows.append(
                        f"**{name}** · {tier} · **${cost:.2f}**\n"
                        f"连续零使用 **{int(decimal(row.get('consecutive_zero_months')))}个月**"
                    )
                else:
                    compact_rows.append(
                        f"**{name}** · {tier} · **${cost:.2f}**\n"
                        f"{_format_credits(row.get('credits'))} / "
                        f"{int(decimal(row.get('capacity'))):,} Credits · "
                        f"{_format_usage_rate(row.get('usage_rate'))} · "
                        f"连续低用 **{int(decimal(row.get('consecutive_low_months')))}个月**"
                    )
            return [{"tag": "markdown", "content": "\n\n".join(compact_rows), "text_size": "normal"}]

        elements: list[dict[str, Any]] = []
        for index, row in enumerate(items):
            if index:
                elements.append({"tag": "hr"})
            elements.append(risk_row(row, kind))
        return elements

    status_text = "部分数据" if status == "PARTIAL" else "数据完整"
    title_suffix = "（部分数据）" if status == "PARTIAL" else ""
    header_tags = []
    if zero_rows:
        header_tags.append({
            "tag": "text_tag", "element_id": "zero_tag", "color": "red",
            "text": {"tag": "plain_text", "content": f"零使用 {len(zero_rows)}"},
        })
    if low_rows:
        header_tags.append({
            "tag": "text_tag", "element_id": "low_tag", "color": "orange",
            "text": {"tag": "plain_text", "content": f"低用 {len(low_rows)}"},
        })

    elements: list[dict[str, Any]] = [
        metric_row(
            metric_column("👥 订阅用户", f"{total_users} 人", "本月纳入分析"),
            metric_column("💰 月度总成本", f"${total_cost:.2f}", "固定套餐成本"),
        ),
        metric_row(
            metric_column(
                "⛔ 零使用", f"{len(zero_rows)} 人 · {format_percent(len(zero_rows), total_users)}",
                f"订阅成本 ${zero_cost:.2f}",
            ),
            metric_column(
                "⚠️ 非零低用量", f"{len(low_rows)} 人 · {format_percent(len(low_rows), total_users)}",
                f"低效容量价值 ${low_idle_value:.2f}",
            ),
        ),
        {
            "tag": "markdown", "text_align": "center", "text_size": "notation",
            "content": (
                f"中用量 **{yellow_count}人（{format_percent(yellow_count, total_users)}）**　·　"
                f"高用量 **{green_count}人（{format_percent(green_count, total_users)}）**"
            ),
        },
    ]

    if zero_rows:
        elements.append({
            "tag": "collapsible_panel", "element_id": "zero_users",
            "expanded": True, "background_color": "panel_bg",
            "vertical_spacing": "4px", "padding": "8px 8px 8px 8px",
            "margin": "4px 0px 0px 0px",
            "header": {
                "title": {
                    "tag": "markdown",
                    "content": f"⛔ **零使用用户 · {len(zero_rows)}人**　订阅成本 ${zero_cost:.2f}",
                },
                "background_color": "neutral_bg", "padding": "8px 8px 8px 8px",
                "icon": {
                    "tag": "standard_icon", "token": "down-small-ccm_outlined",
                    "color": "grey", "size": "16px 16px",
                },
                "icon_position": "right", "icon_expanded_angle": -180,
            },
            "border": {"color": "border_soft", "corner_radius": "6px"},
            "elements": panel_rows(zero_rows, "zero"),
        })
    else:
        elements.append({
            "tag": "markdown", "text_align": "center",
            "content": "**本月无零使用用户**",
        })

    if low_rows:
        elements.append({
            "tag": "collapsible_panel", "element_id": "low_users",
            "expanded": risk_count <= 20, "background_color": "panel_bg",
            "vertical_spacing": "4px", "padding": "8px 8px 8px 8px",
            "margin": "4px 0px 0px 0px",
            "header": {
                "title": {
                    "tag": "markdown",
                    "content": f"⚠️ **非零低用量用户 · {len(low_rows)}人**　低效价值 ${low_idle_value:.2f}",
                },
                "background_color": "neutral_bg", "padding": "8px 8px 8px 8px",
                "icon": {
                    "tag": "standard_icon", "token": "down-small-ccm_outlined",
                    "color": "grey", "size": "16px 16px",
                },
                "icon_position": "right", "icon_expanded_angle": -180,
            },
            "border": {"color": "border_soft", "corner_radius": "6px"},
            "elements": panel_rows(low_rows, "low"),
        })
    else:
        elements.append({
            "tag": "markdown", "text_align": "center",
            "content": "**本月无非零低用量用户**",
        })

    chronic_zero = sum(int(decimal(row.get("consecutive_zero_months"))) >= 2 for row in zero_rows)
    first_zero = sum(int(decimal(row.get("consecutive_zero_months"))) <= 1 for row in zero_rows)
    chronic_low = sum(int(decimal(row.get("consecutive_low_months"))) >= 2 for row in low_rows)
    first_low = sum(int(decimal(row.get("consecutive_low_months"))) <= 1 for row in low_rows)
    suggestions = []
    if chronic_zero:
        suggestions.append(f"- 连续零使用 ≥2个月：**{chronic_zero}人**，建议优先评估回收订阅")
    if first_zero:
        suggestions.append(f"- 首次/单月零使用：**{first_zero}人**，建议提醒确认后续使用计划")
    if chronic_low:
        suggestions.append(f"- 连续低用 ≥2个月：**{chronic_low}人**，建议核查需求或调整套餐")
    if first_low:
        suggestions.append(f"- 首次/单月低用：**{first_low}人**，建议提醒关注后续使用情况")
    if not zero_rows and not low_rows:
        suggestions.append("- 本月没有需要回收或调整的低效订阅")
    elements.append({
        "tag": "column_set", "flex_mode": "stretch", "background_style": "advice_bg",
        "margin": "4px 0px 0px 0px",
        "columns": [{
            "tag": "column", "width": "weighted", "weight": 1,
            "padding": "10px 12px 10px 12px",
            "elements": [{
                "tag": "markdown", "text_size": "normal",
                "content": "💡 **管理建议**\n" + "\n".join(suggestions),
            }],
        }],
    })
    if warning:
        elements.append({
            "tag": "markdown", "text_size": "notation",
            "content": f"**数据说明：** {sanitize(warning)}",
        })
    elements.append({
        "tag": "button", "element_id": "open_report",
        "type": "primary_filled", "size": "large", "width": "fill",
        "margin": "4px 0px 0px 0px",
        "text": {"tag": "plain_text", "content": "📊 查看完整 Excel 报告"},
        "behaviors": [{
            "type": "open_url", "default_url": url,
            "pc_url": url, "ios_url": url, "android_url": url,
        }],
    })

    return {
        "msg_type": "interactive",
        "card": {
            "schema": "2.0",
            "config": {
                "enable_forward": True, "update_multi": True, "width_mode": "fill",
                "style": {
                    "color": {
                        "neutral_bg": {
                            "light_mode": "rgba(245,246,247,0.92)",
                            "dark_mode": "rgba(41,41,41,0.82)",
                        },
                        "panel_bg": {
                            "light_mode": "rgba(250,251,252,0.98)",
                            "dark_mode": "rgba(26,26,26,0.94)",
                        },
                        "border_soft": {
                            "light_mode": "rgba(31,35,41,0.10)",
                            "dark_mode": "rgba(255,255,255,0.12)",
                        },
                        "advice_bg": {
                            "light_mode": "rgba(20,86,240,0.06)",
                            "dark_mode": "rgba(117,164,255,0.10)",
                        },
                    },
                },
                "summary": {
                    "content": (
                        f"{month} Kiro月报：零使用{len(zero_rows)}人，"
                        f"非零低用{len(low_rows)}人，总成本${total_cost:.2f}"
                    ),
                },
            },
            "card_link": {"url": url},
            "header": {
                "template": "blue", "padding": "12px 12px 12px 12px",
                "title": {
                    "tag": "plain_text",
                    "content": f"【{month}】Kiro订阅用户月度用量分析月报{title_suffix}",
                },
                "subtitle": {"tag": "plain_text", "content": f"自然月 · Final · {status_text}"},
                "text_tag_list": header_tags,
            },
            "body": {
                "direction": "vertical", "padding": "12px 12px 12px 12px",
                "vertical_spacing": "12px", "horizontal_align": "left",
                "elements": elements,
            },
        },
    }


def send_feishu(secrets: Any, secret_arn: str, payload: dict[str, Any], timeout: int = 10) -> str | None:
    if not secret_arn:
        return None
    try:
        raw = secrets.get_secret_value(SecretId=secret_arn).get("SecretString", "{}")
        value = json.loads(raw)
        webhook = value.get("webhook") or value.get("url")
        if not webhook:
            raise ValueError("secret JSON requires webhook or url")
        request_payload = dict(payload)
        if value.get("sign_secret"):
            timestamp = str(int(time.time()))
            request_payload.update({"timestamp": timestamp, "sign": feishu_signature(value["sign_secret"], timestamp)})
        request = urllib.request.Request(webhook, data=json.dumps(request_payload, ensure_ascii=False).encode("utf-8"), headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(request, timeout=timeout) as response:
            if response.status >= 300:
                raise RuntimeError("Feishu HTTP request failed")
            result = json.loads(response.read().decode("utf-8"))
            if result.get("code") != 0:
                raise RuntimeError("Feishu API rejected message")
        return None
    except Exception as exc:
        # Do not include exception text: urllib errors can embed the secret webhook URL.
        return f"Feishu notification failed: {type(exc).__name__}"


def send_feishu_channels(secrets: Any, secret_arns: dict[str, str], channels: list[str],
                         payload: dict[str, Any]) -> dict[str, str | None]:
    """Send independently to explicit channels without fallback or secret disclosure."""
    configured = [secret_arns.get(channel, "") for channel in channels]
    if len(channels) > 1 and all(configured) and len(set(configured)) != len(configured):
        return {channel: "Feishu notification configuration error" for channel in channels}
    results: dict[str, str | None] = {}
    for channel in channels:
        secret_arn = secret_arns.get(channel, "")
        if not secret_arn:
            results[channel] = "Feishu notification channel is not configured"
            continue
        results[channel] = send_feishu(secrets, secret_arn, payload)
    return results


def _put(s3: Any, bucket: str, key: str, body: bytes, content_type: str) -> None:
    s3.put_object(Bucket=bucket, Key=key, Body=body, ContentType=content_type, ServerSideEncryption="AES256")


def handler(event: dict[str, Any], context: Any) -> dict[str, Any]:
    event = event or {}
    selection = select_report_month(event)
    month, report_type = selection["month"], selection["report_type"]
    region = os.environ.get("AWS_REGION_NAME") or os.environ.get("AWS_REGION", "us-east-1")
    data_bucket = os.environ["DATA_BUCKET"]
    report_bucket = os.environ["REPORT_BUCKET"]
    output_prefix = os.environ.get("MONTHLY_OUTPUT_PREFIX", "dashboard-reports/public/kiro-monthly")
    database = os.environ.get("GLUE_DATABASE", "kiro_analytics")
    workgroup = os.environ.get("ATHENA_WORKGROUP", "kiro-analytics-workgroup")
    identity_store_id = os.environ["IDENTITY_STORE_ID"]
    clients = {
        "s3": boto3.client("s3", region_name=region), "athena": boto3.client("athena", region_name=region),
        "sso": boto3.client("sso-admin", region_name=region), "identity": boto3.client("identitystore", region_name=region),
        "secrets": boto3.client("secretsmanager", region_name=region),
    }
    usage = usage_rows_by_id(run_athena_query(clients["athena"], build_usage_sql(database, month), workgroup))
    roster, authoritative = load_csv_roster(clients["s3"], data_bucket, os.environ.get("SUBSCRIPTION_CSV_KEY", ""))
    if not authoritative:
        latest_rows = run_athena_query(clients["athena"], build_latest_tiers_sql(database), workgroup)
        observed_tiers = {canonical_user_id(r.get("user_id")): normalize_tier(r.get("latest_tier")) for r in latest_rows if canonical_user_id(r.get("user_id"))}
        roster = resolve_application_roster(clients["sso"], clients["identity"], os.environ.get("KIRO_APPLICATION_ARN", ""), identity_store_id, observed_tiers)
    roster = roster or {}
    event_time = datetime.fromisoformat(str(event.get("time", datetime.now(timezone.utc).isoformat())).replace("Z", "+00:00"))
    current_month = f"{event_time.year:04d}-{event_time.month:02d}"
    include_roster = should_use_current_roster(month, current_month)
    previous, history = load_history(clients["s3"], report_bucket, output_prefix, month)
    identity_map = merge_identity_maps(
        identities_from_history(history),
        load_identity_mapping(clients["s3"], data_bucket, os.environ.get("IDENTITY_MAPPING_KEY", IDENTITY_MAPPING_KEY)),
        identity_projection(roster),
    )
    rows = merge_report_rows(month, usage, roster, include_roster, previous, history, identity_map)
    status, warning = report_status(month)
    if month == current_month:
        status = "PARTIAL"
        warning = (warning + " " if warning else "") + "当前自然月尚未结束，报告为截至运行时可用数据的部分快照。"
    if not include_roster:
        warning = (warning + " " if warning else "") + "历史订阅名册无法重建；本报告仅包含当月使用身份，不将当前名册零使用用户认定为历史订阅。"
    keys = monthly_keys(output_prefix, month, report_type)
    workbook = build_monthly_workbook(rows, month, status, warning)
    detail_csv = rows_to_csv(rows)
    snapshot_csv = subscription_snapshot_csv(roster) if should_write_subscription_snapshot(include_roster) else None
    # Durable monthly writes happen before annual rebuild and notification.
    _put(clients["s3"], report_bucket, keys["workbook"], workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if report_type == "final":
        _put(clients["s3"], report_bucket, keys["alias"], workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    _put(clients["s3"], report_bucket, keys["detail_csv"], detail_csv, "text/csv; charset=utf-8")
    if snapshot_csv is not None:
        _put(clients["s3"], report_bucket, keys["subscription_csv"], snapshot_csv, "text/csv; charset=utf-8")
    annual_rows = list_monthly_snapshots(clients["s3"], report_bucket, output_prefix, month[:4])
    annual = build_annual_workbook(annual_rows, month[:4])
    _put(clients["s3"], report_bucket, keys["annual"], annual, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    url_key = keys["alias"] if report_type == "final" else keys["workbook"]
    url = public_url(report_bucket, region, url_key)
    notification_error = None
    notification_channels: list[str] = []
    notification_results: dict[str, str | None] = {}
    if selection["notify"]:
        notification_channels = select_notification_channels(event)
        payload = build_feishu_card(month, rows, url, status, warning)
        notification_results = send_feishu_channels(
            clients["secrets"],
            {
                "dev": os.environ.get("FEISHU_DEV_SECRET_ARN", ""),
                "prod": os.environ.get("FEISHU_PROD_SECRET_ARN", ""),
            },
            notification_channels,
            payload,
        )
        errors = [f"{channel}: {error}" for channel, error in notification_results.items() if error]
        notification_error = "; ".join(errors) or None
    return {
        "month": month, "report_type": report_type, "status": status, "users": len(rows),
        "report_url": url, "notification_attempted": bool(selection["notify"]),
        "notification_channels": notification_channels, "notification_results": notification_results,
        "notification_error": notification_error, "warning": warning,
    }
